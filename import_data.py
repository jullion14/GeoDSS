"""
GeoDSS — database setup and data import
=======================================

Creates the PostGIS schema and imports every dataset the application needs.
This is the only script required to go from an empty database to a fully
populated one.

SETUP
-----
1. Create the database and enable PostGIS:

       CREATE DATABASE geodss_db;
       \\c geodss_db
       CREATE EXTENSION postgis;

2. Install dependencies:

       pip install psycopg2-binary openpyxl

3. Set the connection details below, or via environment variables:

       PGHOST, PGPORT, PGDATABASE, PGUSER, PGPASSWORD

4. Place the source files in the data/ directory (see DATA_FILES below).

5. Run:

       python import_data.py --dry-run    # validate files, no DB changes
       python import_data.py              # create schema and import

WARNING
-------
This script DROPS and recreates all four tables on every run. All data is
derived from the source files, so re-running is safe and idempotent — but
any manual edits to the tables will be lost.

DATA SOURCES
------------
All from data.gov.sg unless noted:
  - Master Plan 2019 Planning Area Boundary (No Sea)  [URA]
  - Resident Population by Planning Area, Age Group, Sex and Type of
    Dwelling, 2025 (respopagesextod2025e.xlsx)        [SingStat]
  - GP Locations                                      [MOH]
  - Polyclinics                                       [MOH]
  - LTA MRT Station Exit                              [LTA]

All spatial data is published in WGS84 (EPSG:4326); no reprojection needed.
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2
from psycopg2.extras import execute_values

# ---------------------------------------------------------------------------
# CONFIGURATION
# ---------------------------------------------------------------------------

DB_CONFIG = {
    "host":     os.getenv("PGHOST", "localhost"),
    "port":     int(os.getenv("PGPORT", 5432)),
    "dbname":   os.getenv("PGDATABASE", "geodss_db"),
    "user":     os.getenv("PGUSER", "postgres"),
    "password": os.getenv("PGPASSWORD", "Chelseafc14"),  # <-- change this
}

DATA_DIR = "data"

# Required: the import fails if any of these are missing.
DATA_FILES = {
    "planning_areas": "MasterPlan2019PlanningAreaBoundaryNoSea.geojson",
    "population":     "respopagesextod2025e.xlsx",
    "gp":             "GP_Locations.geojson",
    "polyclinic":     "Polyclinics.geojson",
    "transit":        "LTAMRTStationExitGEOJSON.geojson",
}

# Optional: skipped with a note if absent. Bus stops come from the LTA
# DataMall API, which needs a free account key, so a tester can run the
# rest of the setup without one. Generate with:
#     python fetch_bus_data.py
OPTIONAL_FILES = {
    "bus_stops": "bus_stops.geojson",
}

POPULATION_SHEET = "2025(Total)"
POPULATION_YEAR = 2025

# Below this population, derived percentages are unreliable: source figures
# are rounded to the nearest 10 and small cells are reported as "-"
# (nil or negligible), so bands can sum to 0% or exceed 100%. Counts are
# still stored; only the derived percentage fields are set NULL.
MIN_POPULATION_FOR_PCT = 1000


# ---------------------------------------------------------------------------
# SCHEMA
# ---------------------------------------------------------------------------

SCHEMA_SQL = """
DROP TABLE IF EXISTS population         CASCADE;
DROP TABLE IF EXISTS healthcare_facilities CASCADE;
DROP TABLE IF EXISTS transit_exits      CASCADE;
DROP TABLE IF EXISTS bus_stops          CASCADE;
DROP TABLE IF EXISTS planning_areas     CASCADE;

-- Planning areas: the geometric backbone every other table joins against.
CREATE TABLE planning_areas (
    id     SERIAL PRIMARY KEY,
    name   TEXT UNIQUE NOT NULL,
    region TEXT,
    geom   GEOMETRY(MultiPolygon, 4326) NOT NULL
);
CREATE INDEX idx_planning_areas_geom ON planning_areas USING GIST (geom);

-- Resident population by planning area (2025), with age distribution and
-- breakdown by type of dwelling.
CREATE TABLE population (
    planning_area_id INT PRIMARY KEY REFERENCES planning_areas(id),
    year             INT NOT NULL,
    total_population INT,

    -- age distribution, as a percentage of total_population
    pct_below15      NUMERIC(4,1),
    pct_15_24        NUMERIC(4,1),
    pct_25_34        NUMERIC(4,1),
    pct_35_44        NUMERIC(4,1),
    pct_45_54        NUMERIC(4,1),
    pct_55_64        NUMERIC(4,1),
    pct_65andabove   NUMERIC(4,1),
    pct_75andabove   NUMERIC(4,1),

    -- residents by type of dwelling (counts)
    hdb_total        INT,
    hdb_1_2_room     INT,
    hdb_3_room       INT,
    hdb_4_room       INT,
    hdb_5_room_exec  INT,
    condo_other      INT,
    landed           INT,
    dwelling_others  INT,

    -- derived indicators, as a percentage of total_population
    pct_hdb          NUMERIC(4,1),
    pct_hdb_1_2_room NUMERIC(4,1)
);

COMMENT ON TABLE population IS
    'Singapore resident population by planning area, 2025. Figures rounded '
    'to the nearest 10; component rows may not sum exactly to totals. '
    'Percentage fields are NULL where population is too small for them to '
    'be meaningful.';

-- GP clinics and polyclinics in a single table, distinguished by type.
CREATE TABLE healthcare_facilities (
    id               SERIAL PRIMARY KEY,
    name             TEXT NOT NULL,
    facility_type    TEXT NOT NULL CHECK (facility_type IN ('GP', 'Polyclinic')),
    address          TEXT,
    postal_code      TEXT,
    planning_area_id INT REFERENCES planning_areas(id),
    geom             GEOMETRY(Point, 4326) NOT NULL
);
CREATE INDEX idx_healthcare_geom ON healthcare_facilities USING GIST (geom);
CREATE INDEX idx_healthcare_area ON healthcare_facilities (planning_area_id);

-- MRT and LRT station exits.
CREATE TABLE transit_exits (
    id               SERIAL PRIMARY KEY,
    station_name     TEXT NOT NULL,
    exit_code        TEXT,
    planning_area_id INT REFERENCES planning_areas(id),
    geom             GEOMETRY(Point, 4326) NOT NULL
);
CREATE INDEX idx_transit_geom ON transit_exits USING GIST (geom);
CREATE INDEX idx_transit_area ON transit_exits (planning_area_id);

-- Bus stops, from LTA DataMall. service_count is the number of distinct bus
-- services calling at the stop; it discriminates between areas far better
-- than distance does, since the bus network is planned around a roughly
-- 400 m walk and stop proximity therefore varies little across Singapore.
CREATE TABLE bus_stops (
    id               SERIAL PRIMARY KEY,
    bus_stop_code    TEXT UNIQUE NOT NULL,
    road_name        TEXT,
    description      TEXT,
    service_count    INT,
    planning_area_id INT REFERENCES planning_areas(id),
    geom             GEOMETRY(Point, 4326) NOT NULL
);
CREATE INDEX idx_bus_stops_geom ON bus_stops USING GIST (geom);
CREATE INDEX idx_bus_stops_area ON bus_stops (planning_area_id);
"""


# ---------------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------------

def path_for(key):
    name = DATA_FILES.get(key) or OPTIONAL_FILES[key]
    return os.path.join(DATA_DIR, name)


def has_file(key):
    return os.path.isfile(path_for(key))


def check_files():
    """Fail early with a clear message if any required source file is missing."""
    missing = [DATA_FILES[k] for k in DATA_FILES if not has_file(k)]
    if missing:
        print(f"ERROR: missing required files in '{DATA_DIR}/':", file=sys.stderr)
        for f in missing:
            print(f"  - {f}", file=sys.stderr)
        sys.exit(1)

    for key, name in OPTIONAL_FILES.items():
        if not has_file(key):
            print(f"NOTE: {name} not found — {key} will be skipped.")
            if key == "bus_stops":
                print("      Generate it with: python fetch_bus_data.py\n")


def clean(value):
    """
    Some source files use the string 'None' rather than a null for missing
    fields. Normalise both to Python None.
    """
    if value is None:
        return None
    text = str(value).strip()
    return None if text in ("", "None", "-") else text


def load_geojson(key):
    with open(path_for(key), encoding="utf-8") as f:
        return json.load(f)["features"]


# ---------------------------------------------------------------------------
# IMPORT: PLANNING AREAS
# ---------------------------------------------------------------------------

def import_planning_areas(cur):
    features = load_geojson("planning_areas")

    rows = [
        (
            f["properties"]["PLN_AREA_N"].title(),
            (f["properties"].get("REGION_N") or "").title() or None,
            json.dumps(f["geometry"]),
        )
        for f in features
    ]

    execute_values(
        cur,
        """
        INSERT INTO planning_areas (name, region, geom)
        VALUES %s
        """,
        rows,
        template="(%s, %s, ST_Multi(ST_SetSRID(ST_GeomFromGeoJSON(%s), 4326)))",
    )
    print(f"  planning_areas       : {len(rows)} rows")

    cur.execute("SELECT name, id FROM planning_areas")
    return dict(cur.fetchall())


# ---------------------------------------------------------------------------
# IMPORT: POPULATION
# ---------------------------------------------------------------------------

# 5-year source bands mapped onto the 10-year bands used by the application.
AGE_BANDS = {
    "pct_below15":    ["0 - 4", "5 - 9", "10 - 14"],
    "pct_15_24":      ["15 - 19", "20 - 24"],
    "pct_25_34":      ["25 - 29", "30 - 34"],
    "pct_35_44":      ["35 - 39", "40 - 44"],
    "pct_45_54":      ["45 - 49", "50 - 54"],
    "pct_55_64":      ["55 - 59", "60 - 64"],
    "pct_65andabove": ["65 - 69", "70 - 74", "75 - 79", "80 - 84",
                       "85 - 89", "90 & Over"],
    # Additional band: mobility constraints rise sharply here, making this
    # more informative than 65+ for healthcare accessibility.
    "pct_75andabove": ["75 - 79", "80 - 84", "85 - 89", "90 & Over"],
}

DWELLING_COLS = {
    "hdb_total":       "Total HDB^",
    "hdb_1_2_room":    "1- and 2-Room Flats*",
    "hdb_3_room":      "3-Room Flats",
    "hdb_4_room":      "4-Room Flats",
    "hdb_5_room_exec": "5-Room and Executive Flats",
    "condo_other":     "Condominiums and Other Apartments",
    "landed":          "Landed Properties",
    "dwelling_others": "Others",
}

POPULATION_COLS = [
    "planning_area_id", "year", "total_population",
    "pct_below15", "pct_15_24", "pct_25_34", "pct_35_44",
    "pct_45_54", "pct_55_64", "pct_65andabove", "pct_75andabove",
    "hdb_total", "hdb_1_2_room", "hdb_3_room", "hdb_4_room",
    "hdb_5_room_exec", "condo_other", "landed", "dwelling_others",
    "pct_hdb", "pct_hdb_1_2_room",
]


def read_population_xlsx():
    """
    Read planning-area level rows from the source workbook.

    The sheet is a cross-tab of Planning Area x Subzone x Age Group x Type
    of Dwelling. Only planning-area totals (Subzone == 'Total') are used;
    subzone rows are ignored but remain available in the source should a
    finer granularity be added later.
    """
    wb = openpyxl.load_workbook(path_for("population"), read_only=True,
                                data_only=True)
    if POPULATION_SHEET not in wb.sheetnames:
        raise SystemExit(
            f"Sheet '{POPULATION_SHEET}' not found. Available: {wb.sheetnames}"
        )

    lookup, areas, seen = {}, [], set()

    for row in wb[POPULATION_SHEET].iter_rows(min_row=4, values_only=True):
        area, subzone, age, dwelling, value = row[:5]

        # Trailing annotation rows carry text in the first column only.
        if area is None or (subzone is None and age is None):
            break
        if area == "Total" or subzone != "Total":
            continue

        # '-' denotes nil or negligible per the sheet's own footer.
        lookup[(area, age, dwelling)] = value if isinstance(value, (int, float)) else 0

        if area not in seen:
            seen.add(area)
            areas.append(area)

    wb.close()
    return lookup, areas


def build_population_records(lookup, areas):
    def pct(part, whole):
        if not whole or whole < MIN_POPULATION_FOR_PCT:
            return None
        return round(part * 100.0 / whole, 1)

    records = []
    for area in areas:
        total = lookup.get((area, "Total", "Total"), 0)
        rec = {
            "name": area,
            "year": POPULATION_YEAR,
            "total_population": total or None,
        }

        for col, bands in AGE_BANDS.items():
            rec[col] = pct(sum(lookup.get((area, b, "Total"), 0) for b in bands),
                           total)

        for col, label in DWELLING_COLS.items():
            rec[col] = lookup.get((area, "Total", label), 0)

        rec["pct_hdb"] = pct(rec["hdb_total"], total)
        rec["pct_hdb_1_2_room"] = pct(rec["hdb_1_2_room"], total)
        records.append(rec)

    return records


def import_population(cur, area_ids):
    lookup, areas = read_population_xlsx()
    records = build_population_records(lookup, areas)

    rows, unmatched = [], []
    for rec in records:
        area_id = area_ids.get(rec["name"])
        if area_id is None:
            unmatched.append(rec["name"])
            continue
        rec["planning_area_id"] = area_id
        rows.append(tuple(rec.get(c) for c in POPULATION_COLS))

    execute_values(
        cur,
        f"INSERT INTO population ({', '.join(POPULATION_COLS)}) VALUES %s",
        rows,
    )

    populated = [r for r in records if r["total_population"]]
    low = [r for r in populated
           if r["total_population"] < MIN_POPULATION_FOR_PCT]

    print(f"  population           : {len(rows)} rows "
          f"({len(populated)} with residents, "
          f"{len(populated) - len(low)} with computed percentages)")
    if unmatched:
        print(f"    WARNING unmatched planning areas: {unmatched}")
    if low:
        print(f"    percentages NULL below {MIN_POPULATION_FOR_PCT:,} "
              f"residents: {len(low)} areas")


# ---------------------------------------------------------------------------
# IMPORT: POINT LAYERS
# ---------------------------------------------------------------------------

def import_healthcare(cur):
    rows = []

    for f in load_geojson("gp"):
        p = f["properties"]
        rows.append((
            clean(p.get("NAME")) or "Unnamed clinic",
            "GP",
            clean(p.get("ADDRESS")),
            clean(p.get("POSTALCODE")),
            json.dumps(f["geometry"]),
        ))
    gp_count = len(rows)

    for f in load_geojson("polyclinic"):
        p = f["properties"]
        # Polyclinics store address across two fields; either may be absent.
        parts = [clean(p.get("ADDRESSBLOCKHOUSENUMBER")),
                 clean(p.get("ADDRESSSTREETNAME"))]
        rows.append((
            clean(p.get("NAME")) or "Unnamed polyclinic",
            "Polyclinic",
            " ".join(x for x in parts if x) or None,
            clean(p.get("ADDRESSPOSTALCODE")),
            json.dumps(f["geometry"]),
        ))

    execute_values(
        cur,
        """
        INSERT INTO healthcare_facilities
            (name, facility_type, address, postal_code, geom)
        VALUES %s
        """,
        rows,
        template="(%s, %s, %s, %s, ST_SetSRID(ST_GeomFromGeoJSON(%s), 4326))",
    )
    print(f"  healthcare_facilities: {len(rows)} rows "
          f"({gp_count} GPs, {len(rows) - gp_count} polyclinics)")


def import_transit(cur):
    rows = [
        (
            clean(f["properties"].get("STATION_NA")) or "Unnamed station",
            clean(f["properties"].get("EXIT_CODE")),
            json.dumps(f["geometry"]),
        )
        for f in load_geojson("transit")
    ]

    execute_values(
        cur,
        "INSERT INTO transit_exits (station_name, exit_code, geom) VALUES %s",
        rows,
        template="(%s, %s, ST_SetSRID(ST_GeomFromGeoJSON(%s), 4326))",
    )
    print(f"  transit_exits        : {len(rows)} rows")


def import_bus_stops(cur):
    """
    Import bus stops produced by fetch_bus_data.py. Optional: skipped
    silently if the file is absent (check_files already warned).
    """
    if not has_file("bus_stops"):
        return

    rows, with_services = [], 0
    for f in load_geojson("bus_stops"):
        p = f["properties"]
        code = clean(p.get("BusStopCode"))
        if not code:
            continue
        count = p.get("ServiceCount")
        if count is not None:
            with_services += 1
        rows.append((
            code,
            clean(p.get("RoadName")),
            clean(p.get("Description")),
            count,
            json.dumps(f["geometry"]),
        ))

    execute_values(
        cur,
        """
        INSERT INTO bus_stops
            (bus_stop_code, road_name, description, service_count, geom)
        VALUES %s
        ON CONFLICT (bus_stop_code) DO NOTHING
        """,
        rows,
        template="(%s, %s, %s, %s, ST_SetSRID(ST_GeomFromGeoJSON(%s), 4326))",
    )

    note = ("" if with_services == len(rows)
            else f", {len(rows) - with_services} without service counts")
    print(f"  bus_stops            : {len(rows)} rows{note}")


# ---------------------------------------------------------------------------
# SPATIAL JOIN
# ---------------------------------------------------------------------------

def backfill_planning_area_ids(cur):
    """
    Assign each point to the planning area containing it. Pre-computing this
    at import time turns per-area counts into a plain WHERE clause rather
    than a spatial join on every query.
    """
    tables = ["healthcare_facilities", "transit_exits"]
    if has_file("bus_stops"):
        tables.append("bus_stops")

    for table in tables:
        cur.execute(f"""
            UPDATE {table} t
            SET planning_area_id = pa.id
            FROM planning_areas pa
            WHERE ST_Contains(pa.geom, t.geom)
        """)
        matched = cur.rowcount

        cur.execute(f"SELECT COUNT(*) FROM {table} WHERE planning_area_id IS NULL")
        unmatched = cur.fetchone()[0]

        note = ""
        if unmatched:
            # Expected for points on or just outside the "no sea" boundary.
            note = f", {unmatched} outside all boundaries"
        print(f"  {table:<21}: {matched} matched to a planning area{note}")


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def validate_only():
    """Parse every source file without touching the database."""
    print("Validating source files...\n")

    features = load_geojson("planning_areas")
    print(f"  planning areas       : {len(features)} polygons")

    lookup, areas = read_population_xlsx()
    records = build_population_records(lookup, areas)
    populated = [r for r in records if r["total_population"]]
    print(f"  population           : {len(records)} areas, "
          f"{len(populated)} with residents")

    gp = load_geojson("gp")
    poly = load_geojson("polyclinic")
    transit = load_geojson("transit")
    print(f"  healthcare           : {len(gp)} GPs, {len(poly)} polyclinics")
    print(f"  transit exits        : {len(transit)} exits")

    if has_file("bus_stops"):
        stops = load_geojson("bus_stops")
        scored = sum(1 for f in stops
                     if f["properties"].get("ServiceCount") is not None)
        print(f"  bus stops            : {len(stops)} stops, "
              f"{scored} with service counts")
    else:
        print(f"  bus stops            : not present (optional)")

    names = {f["properties"]["PLN_AREA_N"].title() for f in features}
    unmatched = sorted({r["name"] for r in records} - names)
    print(f"\n  name matching        : "
          f"{'all population rows match a planning area' if not unmatched else f'UNMATCHED: {unmatched}'}")

    print("\n  --dry-run: no changes written.")


def main():
    parser = argparse.ArgumentParser(
        description="Create the GeoDSS schema and import all datasets."
    )
    parser.add_argument("--dry-run", action="store_true",
                        help="validate source files without touching the database")
    args = parser.parse_args()

    check_files()

    if args.dry_run:
        validate_only()
        return

    print(f"Connecting to {DB_CONFIG['dbname']} at {DB_CONFIG['host']}...")
    conn = psycopg2.connect(**DB_CONFIG)
    conn.autocommit = False

    try:
        with conn.cursor() as cur:
            cur.execute("SELECT PostGIS_Version()")
            print(f"  PostGIS {cur.fetchone()[0]}\n")

            print("Creating schema (dropping any existing tables)...")
            cur.execute(SCHEMA_SQL)

            print("\nImporting...")
            area_ids = import_planning_areas(cur)
            import_population(cur, area_ids)
            import_healthcare(cur)
            import_transit(cur)
            import_bus_stops(cur)

            print("\nAssigning points to planning areas...")
            backfill_planning_area_ids(cur)

        conn.commit()
        print("\nDone. All data committed.")

    except psycopg2.errors.UndefinedFunction:
        conn.rollback()
        print("\nERROR: PostGIS functions unavailable. Run "
              "'CREATE EXTENSION postgis;' in this database first.",
              file=sys.stderr)
        raise
    except Exception:
        conn.rollback()
        print("\nERROR: import failed, transaction rolled back. "
              "The database is unchanged.", file=sys.stderr)
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    main()